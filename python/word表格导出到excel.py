import docx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from collections import OrderedDict

def extract_tables_from_docx(docx_path, output_excel_path):
    """
    将Word文档中的所有表格导出到Excel的一个sheet中
    
    处理逻辑：
    1. 奇数列（1,3,5...）作为标题/键
    2. 偶数列（2,4,6...）作为值
    3. 所有表格的键值对汇总，相同键的值放在同一列，不同记录占不同行
    4. 支持两列、四列、六列等任意偶数列的表格
    """
    
    # 读取Word文档
    doc = docx.Document(docx_path)
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Word表格汇总"
    
    # 用于存储所有键值对记录
    all_records = []  # 每个元素是一个OrderedDict，代表一条记录
    all_keys_set = set()  # 收集所有出现过的键
    
    # 遍历所有表格
    for table_index, table in enumerate(doc.tables):
        print(f"处理第 {table_index + 1} 个表格，共 {len(table.rows)} 行...")
        
        # 提取表格数据
        table_data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                # 获取单元格文本，处理合并单元格
                cell_text = cell.text.strip()
                # 处理单元格内的换行符
                cell_text = ' '.join(cell_text.split())
                row_data.append(cell_text)
            
            if row_data:
                table_data.append(row_data)
        
        if not table_data:
            print(f"  表格 {table_index + 1} 为空，跳过")
            continue
        
        # 获取表格的列数
        num_cols = max(len(row) for row in table_data)
        print(f"  表格列数: {num_cols}")
        
        # 处理每一行，提取键值对
        for row_idx, row_data in enumerate(table_data):
            record = OrderedDict()
            has_data = False
            
            # 按奇偶列提取键值对
            for col_idx in range(0, len(row_data), 2):
                # 奇数列作为键
                key = row_data[col_idx] if col_idx < len(row_data) else ""
                # 偶数列作为值
                value = row_data[col_idx + 1] if col_idx + 1 < len(row_data) else ""
                
                # 如果键为空，生成默认键名
                if not key:
                    key = f"字段_{table_index+1}_{row_idx+1}_{col_idx//2+1}"
                
                # 清理键名（去除多余空格和特殊字符）
                key = key.strip()
                value = value.strip() if value else ""
                
                record[key] = value
                all_keys_set.add(key)
                
                if value:  # 如果有值，标记为有数据
                    has_data = True
            
            # 只有当这行有实际数据时才添加
            if has_data and record:
                all_records.append(record)
                print(f"    提取记录: {dict(record)}")
    
    if not all_records:
        print("警告：没有提取到任何数据！")
        # 创建一个空文件
        ws.cell(row=1, column=1, value="未提取到数据")
        wb.save(output_excel_path)
        return
    
    # 按照键首次出现的顺序排列标题
    # 这样可以保持一定的顺序性
    ordered_keys = []
    seen_keys = set()
    for record in all_records:
        for key in record.keys():
            if key not in seen_keys:
                ordered_keys.append(key)
                seen_keys.add(key)
    
    print(f"\n共提取到 {len(all_records)} 条记录，{len(ordered_keys)} 个字段")
    print(f"字段列表: {ordered_keys}")
    
    # 设置样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    for col_idx, key in enumerate(ordered_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据行
    for row_idx, record in enumerate(all_records, 2):
        for col_idx, key in enumerate(ordered_keys, 1):
            value = record.get(key, "")  # 如果该记录没有这个键，填空值
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # 自动调整列宽
    for col_idx, key in enumerate(ordered_keys, 1):
        # 计算最大宽度
        max_width = len(key) * 1.5  # 标题宽度（中文字符占2个宽度）
        for row_idx in range(2, len(all_records) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            # 粗略计算中文字符宽度
            width = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_value)
            max_width = max(max_width, width)
        
        # 设置列宽，限制最大宽度
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_width + 2, 50)
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 添加筛选
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(ordered_keys))}{len(all_records) + 1}"
    
    # 保存Excel文件
    wb.save(output_excel_path)
    print(f"\n成功导出到: {output_excel_path}")
    print(f"共 {len(all_records)} 行数据，{len(ordered_keys)} 列")

def main():
    """主函数"""
    import sys
    
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  python extract_tables.py <Word文档路径> [Excel输出路径]")
        print("\n示例:")
        print("  python extract_tables.py input.docx")
        print("  python extract_tables.py input.docx output.xlsx")
        return
    
    docx_path = sys.argv[1]
    
    # 生成输出路径
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base_name = os.path.splitext(docx_path)[0]
        output_path = f"{base_name}_表格汇总.xlsx"
    
    # 检查输入文件是否存在
    if not os.path.exists(docx_path):
        print(f"错误：文件不存在 - {docx_path}")
        return
    
    try:
        extract_tables_from_docx(docx_path, output_path)
    except Exception as e:
        print(f"处理出错: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()