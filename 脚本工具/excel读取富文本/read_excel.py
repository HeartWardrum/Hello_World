import openpyxl
from openpyxl.cell.text import InlineFont
import pandas as pd

def parse_rich_text_cell(cell):
    """
    核心解析函数：提取单个单元格中富文本的加粗片段并转化为 HTML <strong> 标签
    """
    if cell.value is None:
        return ""
        
    # 判断是否为富文本对象 (含有多个文本块)
    if hasattr(cell.value, '__iter__') and not isinstance(cell.value, (str, bytes)):
        html_fragments = []
        
        for text_block in cell.value:
            if isinstance(text_block, str):
                # 纯字符串块，继承单元格整体的加粗状态
                if cell.font and cell.font.bold:
                    html_fragments.append(f"<strong>{text_block}</strong>")
                else:
                    html_fragments.append(text_block)
            else:
                # 独立的富文本块，检查各自的加粗状态
                text = text_block.text
                font = text_block.font
                if font and font.bold:
                    html_fragments.append(f"<strong>{text}</strong>")
                else:
                    html_fragments.append(text)
                    
        return "".join(html_fragments)
    
    else:
        # 普通单元格（整格纯文本、数字等）
        val_str = str(cell.value)
        if cell.font and cell.font.bold:
            return f"<strong>{val_str}</strong>"
        return val_str

def read_all_sheets_with_richtext(file_path):
    """
    读取 Excel 文件中的所有 Sheet 页，并解析其中的富文本加粗标签。
    返回一个字典：{"SheetName": DataFrame}
    """
    # 加载工作簿（切记：处理富文本不能设置 data_only=True）
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # 存储所有 sheet 结果的字典
    all_sheets_dict = {}
    
    # 遍历 Excel 里的每一个 sheet 页
    for sheet_name in wb.sheetnames:
        print(f"正在解析 Sheet: {sheet_name} ...")
        ws = wb[sheet_name]
        
        sheet_data = []
        # 逐行处理
        for row in ws.iter_rows(values_only=False):
            row_data = [parse_rich_text_cell(cell) for cell in row]
            sheet_data.append(row_data)
            
        # 如果当前 sheet 是空的，跳过或给个空 DataFrame
        if not sheet_data:
            all_sheets_dict[sheet_name] = pd.DataFrame()
            continue
            
        # 转换为 DataFrame：默认第一行为表头，其余为数据
        df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
        
        # 将处理好的 DataFrame 存入字典
        all_sheets_dict[sheet_name] = df
        
    print("✨ 所有 Sheet 页解析完毕！")
    return all_sheets_dict

# ==========================================
#                  如何使用
# ==========================================
if __name__ == "__main__":
    # 1. 设置不省略打印，方便你观察完整的 HTML 标签
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_colwidth', None)
    
    # 2. 替换为你的 Excel 文件路径
    excel_path = "你的文件.xlsx" 
    
    # 3. 一次性读取并解析 4 个 sheet
    sheets_data = read_all_sheets_with_richtext(excel_path)
    
    # 4. 查看读取到了哪些 sheet 页
    print("\n成功读取的 Sheet 列表:", list(sheets_data.keys()))
    print("-" * 50)
    
    # 5. 访问具体的某一个 sheet（用你实际的 sheet 名字替换）
    # 假设你的第一个 sheet 叫 'Sheet1'
    first_sheet_name = list(sheets_data.keys())[0] 
    print(f"展示第一个 Sheet ({first_sheet_name}) 的前几行数据：")
    print(sheets_data[first_sheet_name].head())